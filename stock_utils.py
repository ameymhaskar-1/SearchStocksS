import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
import requests

def get_ticker_from_name(company_name):
    """
    Tries to find the NSE ticker for a given company name.
    """
    company_name = company_name.strip()
    if not company_name:
        return None
    
    # 1. If it looks like a ticker (e.g. RELIANCE), just add .NS
    if company_name.isalpha() and len(company_name) <= 10:
        return f"{company_name.upper()}.NS"
    
    # 2. Use yfinance search to find the best match
    try:
        search_url = f"https://query2.finance.yahoo.com/v1/finance/search?q={company_name}"
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(search_url, headers=headers).json()
        for quote in response.get('quotes', []):
            if quote.get('exchange') == 'NSI': # National Stock Exchange of India
                return quote.get('symbol')
    except:
        pass
    
    return None

def fetch_stock_data(tickers):
    data_list = []
    failed_names = []

    for ticker_input in tickers:
        ticker = get_ticker_from_name(ticker_input)
        
        if not ticker:
            failed_names.append(ticker_input)
            continue
            
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            
            # Fetch historical data for 4M to ensure enough data
            hist = stock.history(period="4mo") 
            if hist.empty:
                failed_names.append(ticker_input)
                continue
                
            # Current Data
            cmp = info.get('currentPrice') or hist['Close'].iloc[-1]
            high_52w = info.get('fiftyTwoWeekHigh')
            low_52w = info.get('fiftyTwoWeekLow')
            
            # Market Cap Calculation (Convert to ₹ Cr)
            market_cap = info.get('marketCap')
            market_cap_cr = round(market_cap / 1e7, 2) if market_cap else None
            
            # Historical Prices (Approx 5 trading days = 1 Week, 21 trading days = 1 Month, 63 = 3 Months)
            price_1w = hist['Close'].iloc[-6] if len(hist) >= 6 else hist['Close'].iloc[0]
            price_1m = hist['Close'].iloc[-22] if len(hist) >= 22 else hist['Close'].iloc[0]
            price_3m = hist['Close'].iloc[-64] if len(hist) >= 64 else hist['Close'].iloc[0]

            # Calculations
            ret_1w = ((cmp - price_1w) / price_1w) * 100
            ret_1m = ((cmp - price_1m) / price_1m) * 100
            ret_3m = ((cmp - price_3m) / price_3m) * 100
            dist_high = ((cmp - high_52w) / high_52w) * 100 if high_52w else 0
            dist_low = ((cmp - low_52w) / low_52w) * 100 if low_52w else 0

            # Momentum Logic
            if ret_1m > 0 and ret_3m > 0:
                momentum = "Strong Buying Pressure"
            elif ret_1m < 0 and ret_3m < 0:
                momentum = "Selling Pressure"
            elif ret_1m < 0 and ret_3m > 0:
                momentum = "Profit Booking"
            else:
                momentum = "Recovery"

            data_list.append({
                "Company Name": info.get('longName', ticker_input),
                "NSE Ticker": ticker,
                "CMP": round(cmp, 2),
                "Market Cap (₹ Cr)": market_cap_cr,
                "52W High": round(high_52w, 2) if high_52w else 0,
                "52W Low": round(low_52w, 2) if low_52w else 0,
                "1W Price": round(price_1w, 2),
                "1M Price": round(price_1m, 2),
                "3M Price": round(price_3m, 2),
                "1W Return %": round(ret_1w, 2),
                "1M Return %": round(ret_1m, 2),
                "3M Return %": round(ret_3m, 2),
                "Distance from 52W High %": round(dist_high, 2),
                "Distance from 52W Low %": round(dist_low, 2),
                "Momentum": momentum
            })
        except Exception as e:
            failed_names.append(ticker_input)
            
    return pd.DataFrame(data_list), failed_names

def generate_excel(df):
    wb = Workbook()
    
    # --- Sheet 1: Stock Analysis ---
    ws1 = wb.active
    ws1.title = "Stock Analysis"
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)

    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    # Conditional Formatting for Returns (Row-based loop)
    # Col J (10) = 1W Return, Col K (11) = 1M Return, Col L (12) = 3M Return
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row in range(2, ws1.max_row + 1):
        for col_idx in [10, 11, 12]: # 1W, 1M and 3M Return columns
            cell = ws1.cell(row=row, column=col_idx)
            if cell.value and cell.value > 0:
                cell.fill = green_fill
            elif cell.value and cell.value < 0:
                cell.fill = red_fill

    # Auto-size columns
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws1.column_dimensions[column].width = max_length + 2

    # --- Sheet 2: Pivot Summary ---
    ws2 = wb.create_sheet("Pivot Summary")
    pivot = df.groupby('Momentum').agg({
        'Momentum': 'count',
        '1W Return %': 'mean',
        '1M Return %': 'mean',
        '3M Return %': 'mean'
    }).rename(columns={'Momentum': 'Count'}).reset_index()
    
    for r in dataframe_to_rows(pivot, index=False, header=True):
        ws2.append(r)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = white_font

    # --- Sheet 3: Dashboard ---
    ws3 = wb.create_sheet("Dashboard")
    
    # 1. Momentum Distribution Chart
    pie = PieChart()
    labels = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Momentum Distribution"
    ws3.add_chart(pie, "A1")

    # 2. Top 10 Gainers (1M)
    top_10 = df.nlargest(10, '1M Return %')[['Company Name', '1M Return %']]
    ws3.append(["Top 10 Gainers (1M)"])
    start_row = 20
    for i, r in enumerate(dataframe_to_rows(top_10, index=False, header=True), start=start_row):
        for j, val in enumerate(r, start=1):
            ws3.cell(row=i, column=j).value = val

    bc = BarChart()
    bc.title = "Top 10 Gainers (1M %)"
    bc_data = Reference(ws3, min_col=2, min_row=start_row, max_row=start_row + len(top_10))
    bc_cats = Reference(ws3, min_col=1, min_row=start_row+1, max_row=start_row + len(top_10))
    bc.add_data(bc_data, titles_from_data=True)
    bc.set_categories(bc_cats)
    ws3.add_chart(bc, "I1")

    return wb
